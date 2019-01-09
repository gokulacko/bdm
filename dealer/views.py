from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.core.paginator import Paginator
from django.core import serializers
from django.contrib import messages
import dealer.forms as f
from .forms import BdmForm, DealerForm, ContactForm, ContactEditForm, OutletForm, OutletEditForm, ContactFormOutlet, ContactFormOutletEdit, PriceUploadDealerForm, PriceUploadForm, ContactFormOutletEdit
from .models import Dealer, Bdm, Outlet, Contact, Brand, City, Inventory, Model, Variant, DealerOffer, DealerDiscount, AckodriveDiscount, AckodriveKindOffers, PriceConfig, AckodriveQuote
import dealer.models as m
# from geopy.geocoders import Nominatim
import googlemaps
import datetime
import os
import codecs
from django.conf import settings
from tablib import Dataset
# from .resources import DealerDiscountUploadResource
from django.db.models import Sum

from .filters import InventoryFilter, DealerFilter, PriceFilter
from django.db.models import Q

import openpyxl
from openpyxl import Workbook
import xlwt
import traceback
# from itertools import islice

def index(request):
    
    if request.method == "POST":
        form = BdmForm(request.POST)
        # dealerform = DealerForm(request.POST)
        # contactform = ContactForm(request.POST)
        dealer = Dealer.objects.all()
        if form.is_valid():
            form.save()
            form = BdmForm()
        brand = Brand.objects.all()
        city = City.objects.all()
        context = {
            'form': form,
            # 'dealerform': dealerform,
            # 'contactform': contactform,
            'dealer': dealer,
            'brand':brand,
            'city':city,
            'dealermaster':dealer,
            

        }
        return render(request, 'dealer/index.html', context)
        
    else:
        # if request.POST.get('filter'):
        #     brandpram = request.GET.getlist('brand[]')
        #     citypram = request.GET.getlist('city[]')
        #     print("brand",brandpram)
            
        form = BdmForm()
        # dealerform = DealerForm()
        # contactform = ContactForm()
        # dealer = Dealer.objects.all()
        
        # gmaps = googlemaps.Client(key='')
        # geolocator = Nominatim(timeout= 10)
        # address = "okilipuram, bangalore"
        # add = "9, 17th A Main Rd, Near Sukh Sagar Restaurant, 5th Block, Koramangala, Bengaluru, Karnataka 560095"
        # location = geolocator.geocode(address)
        # print((location.latitude, location.longitude))
        # geocode_result = gmaps.geocode('1600 Amphitheatre Parkway, Mountain View, CA')
        # print(geocode_result)
    brand = Brand.objects.all()
    city = City.objects.all()
    # paginator = Paginator(dealer,10)
    # page = request.GET.get('page')
    # dealer = paginator.get_page(page)
    dealer_list = Dealer.objects.all()
    dealer_filter = DealerFilter(request.GET, queryset=dealer_list)
    dealer_list = dealer_filter.qs
    paginator = Paginator(dealer_list,10)
    page = request.GET.get('page')
    dealer = paginator.get_page(page)
    inventorysum = Inventory.objects.values('dealer').annotate(inventory_sum=Sum('count'))
    
    context = {
                'form': form,
                # 'dealerform': dealerform,
                # 'contactform': contactform,
                'dealer': dealer,
                'brand':brand,
                'city':city,
                'inventorysum':inventorysum,
                'filter': dealer_filter,
                'dealermaster':dealer,
            }
    return render(request, 'dealer/index.html', context)

def addDealer(request):
    
    if request.method == "POST":
        form = DealerForm(request.POST, request.FILES)
        
        if form.is_valid():
            data = form.save()
            if data.sales_outlet:
                Outlet.objects.create(address = data.address, city = data.city, pincode= data.pincode, status=data.status, dealer = data)
            messages.success(request, ' Dealer added successfully')
        else:
            messages.error(request, ' Dealer not added successfully')  
        return redirect('dealer:index')
    else:
        dealerform = DealerForm()
        context = {
                    'dealerform': dealerform,
            }
    return render(request, 'dealer/add_dealer.html', context)
    
def dealer(request, id):
    # searchfiles = ''
    if request.method == "POST":
        # if request.POST.get('month'):
        #     print("month")
        #     today = datetime.date.today()
        #     searchfiles = DealerPriceFile.objects.filter(dealer_id=id, period__month=today.month, period__year=today.year)
        #     if searchfiles:
        #         print("file")
        #         searchfiles = searchfiles[0]
        #     else:
        #         searchfiles = ''
        # else:
        # form = DealerDiscountForm(request.POST, request.FILES)
        inventoryform = f.InventoryForm(request.POST)
        # if form.is_valid():
        #     form.save()
        #     messages.success(request, ' Price upload successfully')
        #     return redirect('index')
        if inventoryform.is_valid():
            inventoryform = inventoryform.save(commit=False)
            dealer_info = Dealer.objects.get(id=id)
            inventoryform.dealer = dealer_info
            inventoryform.save()
            messages.success(request, ' inventory added successfully')
            return redirect('dealer:dealer-view', id=dealer_info.id)
            
    storage = messages.get_messages(request)
    dealer_info = Dealer.objects.get(id=id)
    dealer_contact = Contact.objects.filter(dealer_id = id)
    outlet_info = Outlet.objects.filter(dealer_id = id)
    outlet_contact = Contact.objects.filter(outlet__dealer__id = id)
    today = datetime.date.today()
    # files = DealerPriceFile.objects.filter(dealer_id=id, period__month=today.month)
    inventory = Inventory.objects.filter(dealer_id = id)
    inventoryform = f.InventoryForm()
    # if files:
    #     priceform = DealerDiscountForm(instance=files[0])
    # else:
    #     priceform = DealerDiscountForm()
    context = {
                'dealer_info': dealer_info,
                'dealer_contact': dealer_contact,
                'outlet_info': outlet_info,
                'outlet_contact': outlet_contact,
                'messages':storage,
                # 'priceform':priceform,
                'inventoryform':inventoryform,
                'inventory':inventory
                # 'searchfiles':searchfiles
            }

    return render(request, 'dealer/dealer.html', context)

def dealerDiscount(request):
    if request.method == "POST":
        dataset = request.FILES['file_name']
        dealer_id = request.POST['dealer_name']
        city_id = request.POST['city_name']
        wb = openpyxl.load_workbook(dataset)
        sheets = wb.sheetnames
        for sheet in sheets:
            worksheet = wb[sheet]
            excel_data = list()

            for row in worksheet.iter_rows(row_offset=1):
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                try:
                    try:
                        dealer = Dealer.objects.get(id=dealer_id)
                        variant = Variant.objects.get(name = row_data[2], model__name = row_data[1])
                        city = City.objects.get(id = city_id)
                        
                        discount_record = DealerDiscount.objects.filter(variant__pk=variant.id, dealer__pk=dealer.id, city__pk=city.id, is_latest=True)
                        # print("******discount***********", discount_record)

                        offer_record = DealerOffer.objects.filter(variant__pk=variant.id, dealer__pk=dealer.id, city__pk=city.id, is_latest=True)
                        # print("******offer************", offer_record)

                        if discount_record.exists() and offer_record.exists():
                            print("trying to update******", discount_record['is_latest'])
                            discount_record[0]['is_latest'] = False
                            discount_record[0].save()
                            offer_record[0]['is_latest'] = False
                            offer_record[0].save()
                        if row_data[3] != "None" and row_data[4] != "None":
                            dealerdiscount = DealerDiscount.objects.create(variant=variant, discount=row_data[3],dealer=dealer, city=city, is_latest=True)
                            dealeroffer = DealerOffer.objects.create(variant=variant, offers=row_data[4],dealer=dealer, city=city, is_latest=True)
                            print("Data inserted******")
                            
                    except(Dealer.DoesNotExist, Variant.DoesNotExist, City.DoesNotExist):
                        pass
                except Exception as e:
                    print(sheet)
                    print(e)
                    if len(excel_data)==0:
                        row_data.append("error")
                    else:
                        trace_back = traceback.format_exc()
                        message = str(e)+ " " + str(trace_back)
                        row_data.append(str(e))

                    excel_data.append(row_data)
                    pass
        
        if len(excel_data)>0:
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="Dealer-Users.xls"'

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('Dealer-Users')

            row_num = 0
            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            for data in excel_data:
                for col_num in range(len(data)):
                    ws.write(row_num, col_num, data[col_num], font_style)
                row_num += 1

            wb.save(response)
            messages.error(request, "Some error occured!")
            return response
        else:
            messages.success(request, "File uploaded successfully.")
    dealerform = PriceUploadDealerForm()
    context = {
        'dealerdiscountform': dealerform,
    }
    
    return render(request, 'dealer/dealer_discount.html', context)

def ackodriveDiscount(request):
    if request.method == "POST":
        city_id = request.POST['city_name']
        excel_file = request.FILES["file_name"]
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames

        city = City.objects.get(id = city_id)
        for sheet in sheets:
            worksheet = wb[sheet]
            excel_data = list()
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                try:
                    if row_data[1] != "None" or row_data[2] != "None" or row_data[3] != "None" or row_data[2] != "Variant_Name" or row_data[2] != "Bangalore":
                        # try:
                        variant = Variant.objects.get(name=row_data[2], model__name = row_data[1])
                        discount_record = AckodriveDiscount.objects.filter(variant__pk = variant.id, city__pk=city.id, is_latest=True)

                        kindoffer_record = AckodriveKindOffers.objects.filter(variant__pk = variant.id, city__pk=city.id, is_latest=True)

                        if discount_record.exists() and kindoffer_record.exists():
                            discount_record.is_latest = False
                            discount_record.save()
                            kindoffer_record.is_latest = False
                            kindoffer_record.save()
                        discount = AckodriveDiscount.objects.create(discount = row_data[3], variant = variant, city=city, is_latest=True)
                        kindoffer = AckodriveKindOffers.objects.create(offers = row_data[4], variant = variant, city=city, is_latest=True)

                        print("Data inserted******")
                        # except Variant.DoesNotExist:
                        #     pass
                except Exception as e:
                    print(sheet)
                    print(e)
                    if len(excel_data)==0:
                        row_data.append("error")
                    else:
                        trace_back = traceback.format_exc()
                        message = str(e)+ " " + str(trace_back)
                        row_data.append(str(e))
                    excel_data.append(row_data)
                    pass

        if len(excel_data)>3:
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="Ackodrive-Users.xls"'

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('Ackodrive-Users')
            row_num = 0

            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            for data in excel_data:

                for col_num in range(len(data)):
                    ws.write(row_num, col_num, data[col_num], font_style)
                row_num += 1

            wb.save(response)
            messages.error(request, 'Some error occured')
            return response
        else:
            messages.success(request, 'File uploaded successfully')
        # return redirect('dealer:ackodrive_discount')
    uploadForm = PriceUploadForm()
    context = {
        'uploadform': uploadForm
    }
    return render(request, 'dealer/ackodrive_discount.html', context)

def marketprice(request):
    if request.method == "POST":
        excel_file = request.FILES["file_name"]
        wb = openpyxl.load_workbook(excel_file)
        city_id = request.POST['city_name']
        sheets = wb.sheetnames

        city = City.objects.get(id = city_id)
        for sheet in sheets:
            worksheet = wb[sheet]
            excel_data = list()
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                try:
                    # try:
                    variant = Variant.objects.get(name=row_data[2], model__name = row_data[1])
                    # if variant.exists():
                        # try:
                    pc_record = PriceConfig.objects.filter(variant__pk=variant.id, is_latest=True)
                    if pc_record.exists():
                        pc_record['is_latest'] = False
                        pc_record.save()
                        
                    # except (PriceConfig.DoesNotExist):
                    #     pass
                    marketprice = PriceConfig.objects.create(variant=variant, ex_showroom=row_data[3], registration_amount=row_data[4], insurance_premium=row_data[5], environment_compensation=row_data[6], octroi=row_data[7], depot_charges=row_data[8],rsa_amount=row_data[9],extended_warranty_amount=row_data[10], cash_discount=row_data[11], amc=row_data[12], basic_accessories=row_data[13], number_plate=row_data[14], smart_card=row_data[15], mcd_charges=row_data[16], tax_collected_at_source=row_data[17], road_tax=row_data[18], other_charges=row_data[19], city=city, is_latest=True)
                    print("Data inserted******")
                    messages.success(request, "File uploaded successfully")
                    # except Variant.DoesNotExist:
                    #     pass

                except Exception as e:
                    if len(excel_data)==0:
                        row_data.append("error")
                    else:
                        trace_back = traceback.format_exc()
                        message = str(e)+ " " + str(trace_back)
                        row_data.append(str(e))   
                    excel_data.append(row_data)
                    pass
                
        if len(excel_data)>2:
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="MarketPrice-errors.xls"'

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('MarketPrice-errors')
            row_num = 0
            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            for data in excel_data: 
                for col_num in range(len(data)):
                    ws.write(row_num, col_num, data[col_num], font_style)
                row_num += 1
            wb.save(response)
            messages.error(request, "Some error occured!")
            return response

    uploadForm = PriceUploadForm()
    context = {
        'uploadform': uploadForm
    }
    return render(request, 'dealer/marketprice.html', context)

def download(request, value):
    response = HttpResponse(content_type='application/ms-excel')
    if value.lower()=="marketprice":
        meta = PriceConfig._meta
        field_names = [field.name for field in meta.fields]
        row_num = 0
        # data = PriceConfig.objects.filter(testfield=12).order_by('-id')[0]
        
        response['Content-Disposition'] = 'attachment; filename="MarketPrice.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Ashok Leyland')

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        field_names[0] = 'Model'
        del field_names[-3:]
        print(field_names)

        for col_num in range(len(field_names)):
            ws.write(0, col_num, field_names[col_num].title(), font_style)

        font_style = xlwt.XFStyle()
        font_style.font.bold = False 

        # try:
            # b = PriceConfig.objects.filter(updated_at = PriceConfig.objects.filter().order_by('-created_at')[0].updated_at)
            # data = PriceConfig.objects.latest('created_at')
            # data = PriceConfig.objects.all().values().annotate(max_createdat=Max('created_at'))
        data = PriceConfig.objects.filter(is_latest=True)
        if data.exists():
            for d in data:
                row_num += 1
                ws.write(row_num, 0, d.variant.model.name, font_style)
                ws.write(row_num, 1, d.variant.name, font_style)
                ws.write(row_num, 2, d.ex_showroom, font_style)
                ws.write(row_num, 3, d.registration_amount, font_style)
                ws.write(row_num, 4, d.insurance_premium, font_style)
                ws.write(row_num, 5, d.environment_compensation, font_style)
                ws.write(row_num, 6, d.octroi, font_style)
                ws.write(row_num, 7, d.depot_charges, font_style)
                ws.write(row_num, 8, d.rsa_amount, font_style)
                ws.write(row_num, 9, d.extended_warranty_amount, font_style)
                ws.write(row_num, 10, d.cash_discount, font_style)
                ws.write(row_num, 11, d.amc, font_style)
                ws.write(row_num, 12, d.basic_accessories, font_style)
                ws.write(row_num, 13, d.number_plate, font_style)
                ws.write(row_num, 14, d.smart_card, font_style)
                ws.write(row_num, 15, d.mcd_charges, font_style)
                ws.write(row_num, 16, d.tax_collected_at_source, font_style)
                ws.write(row_num, 17, d.road_tax, font_style)
                ws.write(row_num, 18, d.other_charges, font_style)

        # except (PriceConfig.DoesNotExist):
        #     pass

        wb.save(response)
        return response

    elif value.lower()=="ackodrive" or value.lower()=="dealer":
        field_names = ['Model_Name', 'Variant_Name', 'Cash_Discount', 'Non_Cash_Offer']
        row_num = 0

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Ashok Leyland')

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(field_names)):
            ws.write(0, col_num, field_names[col_num].title(), font_style)

        font_style = xlwt.XFStyle()
        font_style.font.bold = False 

        if value.lower()=="ackodrive":
            response['Content-Disposition'] = 'attachment; filename="Ackodrive-Price.xls"'
            akdiscount = AckodriveDiscount.objects.filter(is_latest=True)
            for dis in akdiscount:
                offer = AckodriveKindOffers.objects.filter(variant=dis.variant, is_latest=True)
                if offer.exists():
                    row_num += 1
                    ws.write(row_num, 0, dis.variant.model.name, font_style)
                    ws.write(row_num, 1, dis.variant.name, font_style)
                    ws.write(row_num, 2, dis.discount, font_style)
                    ws.write(row_num, 3, offer[0].offers, font_style)
        else:
            response['Content-Disposition'] = 'attachment; filename="DealerDiscount.xls"'
            ddrecords = DealerDiscount.objects.filter(is_latest=True)
            for record in ddrecords:
                # print("variant id", record.variant.id)
                dealeroffers = DealerOffer.objects.filter(variant__pk=record.variant.id, is_latest=True)
                # print("*******offes*******", dealeroffers)
                if dealeroffers.exists():
                    dealeroffer = dealeroffers.first()
                    # print("offer data******", dealeroffer.id)
                    row_num += 1
                    ws.write(row_num, 0, record.variant.model.name, font_style)
                    ws.write(row_num, 1, record.variant.name, font_style)
                    ws.write(row_num, 2, record.discount, font_style)
                    ws.write(row_num, 3, dealeroffer.offers, font_style)

        wb.save(response)
        return response

    return render(request, 'dealer/marketprice.html')

@login_required(login_url='/accounts/login/')
def welcome(request):
    return render(request, 'dealer/welcome.html')

def dealerPrice(request):
    
    price_list = PriceConfig.objects.all()
    price_filter = PriceFilter(request.GET, queryset=price_list)
    price_list = price_filter.qs
    
    paginator = Paginator(price_list,10)
    page = request.GET.get('page')
    price = paginator.get_page(page)
    
    ackodiscount = AckodriveDiscount.objects.all()
    ackooffer = AckodriveKindOffers.objects.all()
    dealerdiscount = DealerDiscount.objects.all()
    dealeroffer = m.DealerOffer.objects.all()
    inventory = m.Inventory.objects.all
    context = {
                'price':price,
                'filter':price_filter,
                'ackodiscount':ackodiscount,
                'ackooffer':ackooffer,
                'dealerdiscount':dealerdiscount,
                'dealeroffer':dealeroffer,
                'inventory':inventory,


    #             # 'searchfiles':searchfiles
    }
    return render(request, 'dealer/price.html', context)

def dealerEdit(request, id):
    dealer_info = Dealer.objects.get(id=id)
    if request.method == "POST":
        form = f.DealerEditForm(request.POST, instance=dealer_info)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dealer edited successfully')
            
            return redirect('dealer:dealer-view', id=dealer_info.id)
        else:
            messages.error(request, 'Dealer not edited successfully')
            return redirect('dealer:dealer-view', id=dealer_info.id)

        
        # dealeredit = Dealer.objects.get(id=id)
        # # dealeredit.brand = request.POST.get('brand')
        # dealeredit.dealer_company = request.POST.get('dealer_company')
        # dealeredit.dealership_name = request.POST.get('dealership_name')
        # dealeredit.status = request.POST.get('status')
        # dealeredit.address = request.POST.get('address')
        # dealeredit.city = request.POST.get('city')
        # dealeredit.pincode = request.POST.get('pincode')
        # dealeredit.sales_outlet = request.POST.get('sales_outlet')
        # # geolocator = Nominatim(user_agent="ackodrive", timeout= 3)
        # # location = geolocator.geocode(request.POST.get('address'))
        # # print((location.latitude, location.longitude))
        # dealeredit.save()
        # messages.success(request, 'Dealer edited successfully')
        # return redirect('dealer:dealer-view', id=dealeredit.id)

   
    form = f.DealerEditForm(instance=dealer_info)
    context = {
                'dealer_info': dealer_info, 
                'form':form,
            } 
    return render(request, 'dealer/dealer_edit.html', context)

def addOutlet(request, id):
    # outlet_info = Outlet.objects.get(id=id)
    dealer_id = id
    dealers = Dealer.objects.get(id=dealer_id)
    if request.method == "POST":
        form = OutletForm(request.POST)
        
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Outlet added successfully')
            # dealer_id = outlet_info.dealer.id
            return redirect('dealer:dealer-view', id=dealer_id)
        else:
            messages.error(request, 'Outlet not added successfully')
    else:
        form = OutletForm(initial={'dealer': dealers})  
    context = {
                'dealer_id': dealer_id,
                'outletform': form,
            } 
    return render(request, 'dealer/add_outlet.html', context)

def outletEdit(request, id):
    outlet_info = Outlet.objects.get(id=id)
    dealer_id = outlet_info.dealer.id
    if request.method == "POST":
        form = OutletEditForm(request.POST, instance=outlet_info)
        if form.is_valid():
            form.save()
            messages.success(request, 'Outlet edited successfully')
            dealer_id = outlet_info.dealer.id
            return redirect('dealer:dealer-view', id=dealer_id)
        else:
            messages.error(request, 'Outlet not edited successfully')

    form = OutletEditForm(instance=outlet_info) 
    context = {
                'outlet_info': outlet_info,
                'form': form,
                'dealer_id': dealer_id,
            }   
    return render(request, 'dealer/outlet_edit.html', context)

def contactEdit(request, id):
    storage = messages.get_messages(request)
    contact = Contact.objects.get(id=id)
    if contact.dealer:
        dealer_id = contact.dealer.id
    else:
        dealer_id = contact.outlet.dealer.id
    if request.method == "POST":
        
        contactform = ContactEditForm(request.POST, instance=contact)
        if contactform.is_valid():
            contactform.save()
            messages.success(request, 'Contact edited successfully')
        else:
            messages.error(request, 'Contact not edited successfully')
            
        # print(form['active'].value())
        # contact = Contact.objects.get(id=id)
        # contact.name = request.POST.get('name')
        # contact.designation = request.POST.get('designation')
        # contact.email = request.POST.get('email')
        # contact.contact_no_1 = request.POST.get('contact_no_1')
        # contact.contact_no_2 = request.POST.get('contact_no_2')   
        # contact.active = form['active'].value()
        # contact.save()
    contactform = ContactEditForm(instance=contact)
    context = {
                'contact': contact,
                'contactform': contactform,
                'dealer_id': dealer_id,
                'messages':storage, 
            }   
    return render(request, 'dealer/contact_edit.html', context)

def outletContactEdit(request, id):
    storage = messages.get_messages(request)
    contact = Contact.objects.get(id=id)
    dealer_id = contact.outlet.dealer.id
    if request.method == "POST":
        
        contactform = ContactFormOutletEdit(request.POST, instance=contact)
        if contactform.is_valid():
            contactform.save()
            messages.success(request, 'Contact edited successfully')
        else:
            messages.error(request, 'Contact not edited successfully')
            
        # print(form['active'].value())
        # contact = Contact.objects.get(id=id)
        # contact.name = request.POST.get('name')
        # contact.designation = request.POST.get('designation')
        # contact.email = request.POST.get('email')
        # contact.contact_no_1 = request.POST.get('contact_no_1')
        # contact.contact_no_2 = request.POST.get('contact_no_2')   
        # contact.active = form['active'].value()
        # contact.save()
    contactform = ContactFormOutletEdit(instance=contact)
    context = {
                'contact': contact,
                'contactform': contactform,
                'dealer_id': dealer_id,
                'messages':storage, 
            }   
    return render(request, 'dealer/contact_edit.html', context)

def addDealerContact(request, id):
    dealer_id = id
    dealers = Dealer.objects.get(id=dealer_id)
    
    if request.method == "POST":
        contactform = ContactForm(request.POST)
        if contactform.is_valid():
            contactform.save()
            messages.success(request, 'Contact added successfully')
            return redirect('dealer:dealer-view', id=dealer_id)
        else:
            messages.error(request, 'Contact not added successfully')
            return redirect('dealer:dealer-view', id=dealer_id)
            
    contactform = ContactForm(initial={'dealer': dealers, 'type': "dealer"})
    context = {
                
                'contactform': contactform,
                'dealer_id': dealer_id,
                
            } 
    
    return render(request, 'dealer/add_dealer_contact.html', context )

def deleteOutlet(request,id):
    outlet=Outlet.objects.get(id=id)
    dealer_id = outlet.dealer.id
    Contact.objects.filter(outlet_id=id).delete()
    outlet.delete()
    messages.success(request, 'Outlet deleted successfully')
    return redirect('dealer:dealer-view', id=dealer_id)

def deleteDealer(request,id):
    dealer=Dealer.objects.get(id=id)
    outlet=Outlet.objects.filter(dealer_id=id)
    Contact.objects.filter(dealer_id=id).delete()
    for out in outlet:
        Contact.objects.filter(outlet_id=out.id).delete()
   
    outlet.delete()
    dealer.delete()
    messages.success(request, 'Dealer deleted successfully')
    return HttpResponseRedirect('/workspace')
    
def deleteContact(request,id):
    contact=Contact.objects.get(id=id)
    if contact.dealer:
        dealer_id = contact.dealer.id
    else:
        dealer_id = contact.outlet.dealer.id
    contact.delete()
    messages.success(request, 'Contact deleted successfully')
    return redirect('dealer:dealer-view', id=dealer_id)

def addOutletContact(request, id):
    outlet_id = id
    outlet=Outlet.objects.get(id=outlet_id)
    dealer_id = outlet.dealer.id
    
    if request.method == "POST":
        contactform = ContactFormOutlet(request.POST)
        if contactform.is_valid():
            contactform.save()
            messages.success(request, 'Contact added successfully')
            return redirect('dealer:dealer-view', id=dealer_id)
        else:
            messages.error(request, 'Contact not added successfully')
            
    contactform = ContactFormOutlet(initial={'outlet': outlet, 'type': "outlet"})
    context = {
                
                'contactform': contactform,
                'outlet_id': outlet_id,
                'dealer_id':dealer_id,
                'messages':messages,
                
            } 
    return render(request, 'dealer/add_outlet_contact.html', context )

def dealerPriceFileDownload(request, id):
    # storage = messages.get_messages(request)
    # dealer_info = Dealer.objects.get(id=id)
    # dealer_contact = Contact.objects.filter(dealer_id = id)
    # outlet_info = Outlet.objects.filter(dealer_id = id)
    # outlet_contact = Contact.objects.filter(outlet__dealer__id = id)

    # price_file = DealerPriceFile.objects.get(id=id)

    today = datetime.date.today()
    # searchfiles = DealerPriceFile.objects.filter(dealer_id=id, period__month=today.month, period__year=today.year)
    # print(searchfiles[0].file)

    # file_name = os.path.basename(searchfiles[0].file)
    # file_name = searchfiles[0].file.name
    print(file_name)
    # file_path = os.path.join(settings.MEDIA_ROOT, './dealerprice', file_name)
    # print(file_path)
    if isinstance(file_name, str):
        myStr = file_name.encode('utf-8') 
    else:
        myStr = file_name

    with open(myStr,'r') as f:
        response = HttpResponse(f.read(), content_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = 'inline;filename=' + myStr
    return response
    # return HttpResponse("here we go!")

    # if file_name:
    #     with open(searchfiles[0].file, 'r') as fh:
    #         # response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
    #         # response['Content-Disposition'] = 'inline; filename=' + os.path.basename(searchfiles[0].file)
    #         return HttpResponse("here we go!")
    #     raise Http404
    # else:
    #     return HttpResponse("in else case")

    # file_path = os.path.join(settings.MEDIA_ROOT, price_file.file)
    # if os.path.exists(price_file.file):
    #     with open(price_file.file, 'rb') as fh:
    #         response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
    #         response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
    #         return response
    # raise Http404

    # DealerPriceFile
    # context = {
    #     'dealer_info': dealer_info,
    #     'dealer_contact': dealer_contact,
    #     'outlet_info': outlet_info,
    #     'outlet_contact': outlet_contact,
    #     'messages':storage,
    # }
    # return render(request, 'dealer/dealer.html', context)

def inventory(request):
    # if request.method == "POST":
    #     if request.POST.get('filter'):
    #         modelpram = request.POST.get('model')
    #         variantpram = request.POST.get('variant')
    #         brandpram = request.POST.get('brand')
    #         # statuspram = request.POST.get('status')
    #         namesearch = request.POST.get('namesearch')
    #         if not brandpram:
    #             brandpram = ""
    #         if not variantpram:
    #             variantpram = ""
    #         if not modelpram:
    #             modelpram = ""
    #         # if not statuspram:
    #         #     statuspram = ""
    #         if not namesearch:
    #             namesearch = ""
    #         inventory = Inventory.objects.filter(variant__name__icontains=variantpram,
    #             dealer__brand__name__icontains = brandpram,
    #             variant__model__name__icontains = modelpram,
    #             dealer__dealership_name__icontains = namesearch

    #             )
    #         brand = Brand.objects.all()
    #         # city = City.objects.all()
    #         model = Model.objects.all()
    #         variant = Variant.objects.all()
            
    #         paginator = Paginator(inventory,10)
    #         page = request.GET.get('page')
    #         inventory = paginator.get_page(page)
    #         context = {
                
    #             'inventory': inventory,
    #             'brand':brand,
    #             'model':model,
    #             'variant':variant,
    #             # 'city':city,
    #             'brandpram':brandpram,
    #             'modelpram':modelpram,
    #             'variantpram':variantpram,
    #             # 'statuspram':statuspram,
    #             'namesearch':namesearch,

    #         }
    #         return render(request, 'dealer/inventory.html', context)
    inventory = Inventory.objects.all()
    brand = Brand.objects.all()
    # city = City.objects.all()
    model = Model.objects.all()
    variant = Variant.objects.all()
    
    
    
    inventory_filter = InventoryFilter(request.GET, queryset=inventory)
    inventory_list = inventory_filter.qs
    paginator = Paginator(inventory_list,10)
    page = request.GET.get('page')
    inventory = paginator.get_page(page)
    
    context = {
        
        'inventory': inventory,
        'brand':brand,
        'model':model,
        'variant':variant,
        # 'city':city,
        'filter': inventory_filter,

    }

    return render(request, 'dealer/inventory.html', context)

def inventoryEdit(request, id):
    inventory = Inventory.objects.get(id=id)
    
    if request.method == "POST":
        form = f.InventoryForm(request.POST, instance=inventory)
        if form.is_valid():
            form.save()
            messages.success(request, 'Outlet edited successfully')
            
            return redirect('dealer:inventory-edit', id=inventory.id)
        else:
            messages.error(request, 'Outlet not edited successfully')

    form = f.InventoryForm(instance=inventory) 
    context = {
                'inventory': inventory,
                'form': form,
                
            }   


    return render(request, 'dealer/inventory_edit.html', context)

def deleteInventory(request, id):
    inventory=Inventory.objects.get(id=id)
    # if contact.dealer:
    #     dealer_id = contact.dealer.id
    # else:
    #     dealer_id = contact.outlet.dealer.id
    inventory.delete()
    messages.success(request, 'Inventory deleted successfully')
    return redirect('dealer:inventory')

def discount(request):

    return render(request, 'dealer/discount.html')


def priceDetails(request, id):

    price = PriceConfig.objects.get(id = id)
    variantid = price.variant.id
    ackodiscount = AckodriveDiscount.objects.filter(variant_id = variantid)
    ackooffer = AckodriveKindOffers.objects.filter(variant_id = variantid)
    dealerdiscount = DealerDiscount.objects.filter(variant_id = variantid)
    dealeroffer = m.DealerOffer.objects.filter(variant_id = variantid)
    print(ackodiscount)
    context = {
        'price':price,
        'ackodiscount':ackodiscount,
        'ackooffer':ackooffer,
        'dealerdiscount':dealerdiscount,
        'dealeroffer':dealeroffer

    }


    return render(request, 'dealer/pricedetails.html', context)